import io
import fitz
import pandas as pd
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from utils.logger import Logger

logger = Logger(__name__)


class DataHandler:
    @staticmethod
    def extract_from_excel(file_bytes: bytes) -> Dict[str, float]:
        """Extract account names and values from Excel file."""
        try:
            excel_file = io.BytesIO(file_bytes)

            # Try with pandas first for better format detection
            try:
                df = pd.read_excel(excel_file, sheet_name=None)  # Read all sheets
                data_map = {}

                # Try to find financial data in any sheet
                for sheet_name, sheet_df in df.items():
                    # Skip sheets with very little data
                    if len(sheet_df) < 5:
                        continue

                    # Look for year columns (like 2024.0, 2025.0) or numeric headers
                    potential_value_cols = []
                    for col in sheet_df.columns:
                        # Check if column name looks like a year or is numeric
                        col_str = str(col)
                        if col_str.replace(".", "").replace("-", "").isdigit():
                            potential_value_cols.append(col)

                    # If we found year columns, extract data
                    if potential_value_cols:
                        # Get the first column as account names
                        account_col = sheet_df.columns[0]
                        # Use the first numeric column (usually most recent year)
                        value_col = potential_value_cols[0]

                        for idx, row in sheet_df.iterrows():
                            account_name = row[account_col]
                            value = row[value_col]

                            if pd.notna(account_name) and pd.notna(value):
                                # Skip header-like rows
                                if str(account_name).lower() in [
                                    "particulars",
                                    "account",
                                    "description",
                                    "nan",
                                ]:
                                    continue

                                # Try to convert value to float
                                try:
                                    float_value = float(value)
                                    if float_value != 0:  # Skip zero values
                                        data_map[str(account_name).strip()] = (
                                            float_value
                                        )
                                except (ValueError, TypeError):
                                    continue

                # If pandas method found data, return it
                if data_map:
                    logger.info(
                        f"Extracted {len(data_map)} data points from Excel (pandas method)"
                    )
                    return data_map

            except Exception as e:
                logger.warning(f"Pandas extraction failed, trying openpyxl: {e}")

            # Fallback to original openpyxl method for simple two-column format
            excel_file.seek(0)  # Reset file pointer
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            data_map = {}
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                if len(row) >= 2:
                    account_name = row[0].value
                    value = row[1].value

                    if account_name and isinstance(value, (int, float)):
                        account_name_str = str(account_name).strip()
                        if value != 0:  # Skip zero values
                            data_map[account_name_str] = float(value)

            logger.info(
                f"Extracted {len(data_map)} data points from Excel (openpyxl method)"
            )
            return data_map

        except Exception as e:
            logger.error(f"Failed to extract Excel data: {str(e)}")
            raise

    @staticmethod
    def extract_from_pdf(file_bytes: bytes) -> Dict[str, float]:
        """Extract account names and values from PDF table/text."""
        try:
            pdf_document = fitz.open(stream=file_bytes, filetype="pdf")
            data_map = {}

            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]
                text = page.get_text()

                lines = text.split("\n")
                for line in lines:
                    parts = line.split()
                    if len(parts) >= 2:
                        try:
                            value = float(parts[-1].replace(",", ""))
                            account_name = " ".join(parts[:-1]).strip()
                            if account_name:
                                data_map[account_name] = value
                        except ValueError:
                            continue

            pdf_document.close()
            logger.info(f"Extracted {len(data_map)} data points from PDF")
            return data_map

        except Exception as e:
            logger.error(f"Failed to extract PDF data: {str(e)}")
            raise

    @staticmethod
    def normalize_account_name(name: str) -> str:
        """Normalize account name for matching."""
        return name.lower().strip()

    @staticmethod
    def format_number(value: float) -> str:
        """Format number with commas and appropriate decimals."""
        if isinstance(value, float) and value.is_integer():
            return f"{int(value):,}"
        return f"{value:,.2f}".rstrip("0").rstrip(".")
